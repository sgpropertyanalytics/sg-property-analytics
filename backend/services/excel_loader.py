"""
Excel Data Loader - Source-of-Truth Import for New Launches

This service loads curated property data from Excel files.
Excel is treated as the source-of-truth; scrapers only suggest updates.

Usage:
    from services.excel_loader import load_new_launches_excel

    stats = load_new_launches_excel(
        file_path='data/new_launches_2026.xlsx',
        db_session=db.session,
        dry_run=False
    )

Schema validation ensures data quality before import.
"""
import os
from datetime import datetime
from typing import Dict, Any, List, Optional
from decimal import Decimal


# Excel reading - openpyxl for .xlsx
OPENPYXL_AVAILABLE = False
try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    pass

# Also support pandas for convenience
PANDAS_AVAILABLE = False
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    pass


# =============================================================================
# SCHEMA DEFINITION
# =============================================================================

REQUIRED_COLUMNS = [
    'project_name',
    'developer',
    'district',
    'market_segment',
    'total_units',
    'psf_min',
    'tenure',
    'source',
    'confidence',
    'last_updated',
]

OPTIONAL_COLUMNS = [
    'psf_max',
    'launch_date',
    'top_date',
    'address',
    'planning_area',
    'notes',
    'land_bid_psf',
]

VALID_MARKET_SEGMENTS = {'CCR', 'RCR', 'OCR'}
VALID_CONFIDENCE_LEVELS = {'high', 'medium', 'low'}
VALID_TENURES = {'Freehold', '99-year', '999-year'}


class ValidationError(Exception):
    """Raised when Excel data fails validation."""
    pass


# =============================================================================
# VALIDATION
# =============================================================================

def validate_row(row: Dict[str, Any], row_num: int) -> List[str]:
    """
    Validate a single row of data.
    Returns list of error messages (empty if valid).
    """
    errors = []

    # Check required fields
    for col in REQUIRED_COLUMNS:
        if col not in row or row[col] is None or str(row[col]).strip() == '':
            errors.append(f"Row {row_num}: Missing required field '{col}'")

    if errors:
        return errors  # Stop early if required fields missing

    # Validate project_name
    name = str(row.get('project_name', '')).strip()
    if len(name) < 3:
        errors.append(f"Row {row_num}: project_name too short (min 3 chars)")
    if len(name) > 100:
        errors.append(f"Row {row_num}: project_name too long (max 100 chars)")

    # Validate district format (D01-D28)
    district = str(row.get('district', '')).strip().upper()
    if not district.startswith('D') or not district[1:].isdigit():
        errors.append(f"Row {row_num}: district must be D01-D28 format, got '{district}'")
    else:
        d_num = int(district[1:])
        if d_num < 1 or d_num > 28:
            errors.append(f"Row {row_num}: district must be D01-D28, got '{district}'")

    # Validate market_segment
    segment = str(row.get('market_segment', '')).strip().upper()
    if segment not in VALID_MARKET_SEGMENTS:
        errors.append(f"Row {row_num}: market_segment must be CCR/RCR/OCR, got '{segment}'")

    # Validate total_units
    try:
        units = int(row.get('total_units', 0))
        if units < 1:
            errors.append(f"Row {row_num}: total_units must be positive, got {units}")
        if units > 5000:
            errors.append(f"Row {row_num}: total_units suspiciously high ({units}), please verify")
    except (ValueError, TypeError):
        errors.append(f"Row {row_num}: total_units must be an integer")

    # Validate PSF
    try:
        psf_min = float(row.get('psf_min', 0))
        if psf_min < 500 or psf_min > 10000:
            errors.append(f"Row {row_num}: psf_min out of range (500-10000), got {psf_min}")
    except (ValueError, TypeError):
        errors.append(f"Row {row_num}: psf_min must be a number")

    if row.get('psf_max'):
        try:
            psf_max = float(row.get('psf_max', 0))
            if psf_max < psf_min:
                errors.append(f"Row {row_num}: psf_max ({psf_max}) cannot be less than psf_min ({psf_min})")
        except (ValueError, TypeError):
            errors.append(f"Row {row_num}: psf_max must be a number")

    # Validate tenure
    tenure = str(row.get('tenure', '')).strip()
    if tenure not in VALID_TENURES:
        errors.append(f"Row {row_num}: tenure must be Freehold/99-year/999-year, got '{tenure}'")

    # Validate confidence
    confidence = str(row.get('confidence', '')).strip().lower()
    if confidence not in VALID_CONFIDENCE_LEVELS:
        errors.append(f"Row {row_num}: confidence must be high/medium/low, got '{confidence}'")

    # Validate source (just ensure it's not empty)
    source = str(row.get('source', '')).strip()
    if len(source) < 2:
        errors.append(f"Row {row_num}: source is required")

    return errors


def validate_excel_data(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    Validate all rows in the Excel file.
    Returns validation report with errors and warnings.
    """
    report = {
        'valid': True,
        'total_rows': len(rows),
        'valid_rows': 0,
        'errors': [],
        'warnings': [],
        'project_names': set(),
    }

    for i, row in enumerate(rows, start=2):  # Row 1 is header
        row_errors = validate_row(row, i)

        if row_errors:
            report['errors'].extend(row_errors)
            report['valid'] = False
        else:
            report['valid_rows'] += 1

            # Check for duplicate project names
            name = str(row.get('project_name', '')).strip().lower()
            if name in report['project_names']:
                report['warnings'].append(f"Row {i}: Duplicate project_name '{row['project_name']}'")
            report['project_names'].add(name)

    return report


# =============================================================================
# EXCEL READING
# =============================================================================

def read_excel_file(file_path: str) -> List[Dict[str, Any]]:
    """
    Read Excel or CSV file and return list of row dictionaries.
    Uses pandas if available, falls back to openpyxl (Excel) or csv (CSV).
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")

    # Handle CSV files
    if file_path.lower().endswith('.csv'):
        return _read_csv(file_path)

    if PANDAS_AVAILABLE:
        return _read_with_pandas(file_path)
    elif OPENPYXL_AVAILABLE:
        return _read_with_openpyxl(file_path)
    else:
        raise ImportError("Neither pandas nor openpyxl installed. Install with: pip install pandas openpyxl")


def _read_csv(file_path: str) -> List[Dict[str, Any]]:
    """Read CSV file using stdlib csv module."""
    import csv

    rows = []
    with open(file_path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            # Normalize column names
            normalized = {}
            for k, v in row.items():
                key = str(k).strip().lower().replace(' ', '_')
                normalized[key] = v if v else None
            rows.append(normalized)

    return rows


def _read_with_pandas(file_path: str) -> List[Dict[str, Any]]:
    """Read Excel using pandas."""
    import pandas as pd

    df = pd.read_excel(file_path, engine='openpyxl')

    # Normalize column names (lowercase, strip whitespace)
    df.columns = [str(c).strip().lower().replace(' ', '_') for c in df.columns]

    # Convert to list of dicts
    rows = df.to_dict('records')

    return rows


def _read_with_openpyxl(file_path: str) -> List[Dict[str, Any]]:
    """Read Excel using openpyxl directly."""
    wb = load_workbook(file_path, read_only=True)
    ws = wb.active

    rows = []
    headers = None

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            # First row is header
            headers = [str(c).strip().lower().replace(' ', '_') if c else f'col_{j}'
                      for j, c in enumerate(row)]
            continue

        if not any(row):  # Skip empty rows
            continue

        row_dict = {}
        for j, value in enumerate(row):
            if j < len(headers):
                row_dict[headers[j]] = value
        rows.append(row_dict)

    wb.close()
    return rows


# =============================================================================
# DATABASE IMPORT
# =============================================================================

def load_new_launches_excel(
    file_path: str,
    db_session=None,
    dry_run: bool = False,
    reset: bool = False,
    year: int = 2026
) -> Dict[str, Any]:
    """
    Load new launches from Excel file into database.

    Args:
        file_path: Path to Excel file
        db_session: SQLAlchemy session
        dry_run: If True, validate only without saving
        reset: If True, delete existing records for the year first
        year: Launch year (default 2026)

    Returns:
        Import statistics and any errors
    """
    from models.new_launch import NewLaunch
    from models.database import db

    if db_session is None:
        db_session = db.session

    stats = {
        'file': file_path,
        'dry_run': dry_run,
        'rows_read': 0,
        'validation_errors': [],
        'inserted': 0,
        'updated': 0,
        'skipped': 0,
        'errors': [],
    }

    print(f"\n{'='*60}")
    print(f"Excel Import: {file_path}")
    print(f"{'='*60}")
    print(f"Mode: {'DRY RUN' if dry_run else 'LIVE'}")
    print(f"Reset: {'Yes' if reset else 'No'}")
    print(f"Year: {year}")
    print(f"{'='*60}\n")

    # Read Excel file
    try:
        rows = read_excel_file(file_path)
        stats['rows_read'] = len(rows)
        print(f"Read {len(rows)} rows from Excel")
    except Exception as e:
        stats['errors'].append(f"Failed to read Excel: {e}")
        print(f"Error reading Excel: {e}")
        return stats

    # Validate data
    print("\nValidating data...")
    validation = validate_excel_data(rows)
    stats['validation_errors'] = validation['errors']

    if validation['errors']:
        print(f"\n{len(validation['errors'])} validation errors found:")
        for error in validation['errors'][:20]:
            print(f"  - {error}")
        if len(validation['errors']) > 20:
            print(f"  ... and {len(validation['errors']) - 20} more")

        if not dry_run:
            print("\nImport aborted due to validation errors. Fix errors and retry.")
            return stats

    if validation['warnings']:
        print(f"\n{len(validation['warnings'])} warnings:")
        for warning in validation['warnings']:
            print(f"  - {warning}")

    print(f"\nValidation: {validation['valid_rows']}/{validation['total_rows']} rows valid")

    if dry_run:
        print("\nDRY RUN complete. No changes made.")
        return stats

    # Reset existing data if requested
    if reset:
        deleted = db_session.query(NewLaunch).filter(
            NewLaunch.launch_year == year
        ).delete()
        db_session.commit()
        print(f"\nDeleted {deleted} existing records for {year}")

    # Import valid rows
    print("\nImporting data...")
    existing_names = {
        r[0].lower(): r[1] for r in
        db_session.query(NewLaunch.project_name, NewLaunch.id).filter(
            NewLaunch.launch_year == year
        ).all()
    }

    for i, row in enumerate(rows, start=2):
        row_errors = validate_row(row, i)
        if row_errors:
            stats['skipped'] += 1
            continue

        project_name = str(row['project_name']).strip()
        name_lower = project_name.lower()

        try:
            if name_lower in existing_names:
                # Update existing
                existing = db_session.query(NewLaunch).get(existing_names[name_lower])
                if existing:
                    _update_from_excel(existing, row)
                    stats['updated'] += 1
                    print(f"  Updated: {project_name}")
            else:
                # Insert new
                new_launch = _create_from_excel(row, year)
                db_session.add(new_launch)
                stats['inserted'] += 1
                print(f"  Inserted: {project_name}")

            db_session.commit()

        except Exception as e:
            db_session.rollback()
            stats['errors'].append(f"{project_name}: {str(e)}")
            print(f"  Error: {project_name} - {e}")

    print(f"\n{'='*60}")
    print("Import Complete")
    print(f"{'='*60}")
    print(f"Inserted: {stats['inserted']}")
    print(f"Updated: {stats['updated']}")
    print(f"Skipped: {stats['skipped']}")
    if stats['errors']:
        print(f"Errors: {len(stats['errors'])}")

    return stats


def _create_from_excel(row: Dict[str, Any], year: int):
    """Create a NewLaunch record from Excel row."""
    from models.new_launch import NewLaunch

    district = str(row['district']).strip().upper()
    if not district.startswith('D'):
        district = f"D{district.zfill(2)}"

    return NewLaunch(
        project_name=str(row['project_name']).strip(),
        developer=str(row.get('developer', '')).strip() or None,
        district=district,
        planning_area=str(row.get('planning_area', '')).strip() or None,
        market_segment=str(row.get('market_segment', '')).strip().upper() or None,
        address=str(row.get('address', '')).strip() or None,
        total_units=int(row.get('total_units', 0)) or None,
        indicative_psf_low=float(row.get('psf_min', 0)) or None,
        indicative_psf_high=float(row.get('psf_max') or row.get('psf_min', 0)) or None,
        tenure=str(row.get('tenure', '')).strip() or None,
        launch_year=year,
        property_type='Condominium',
        land_bid_psf=float(row.get('land_bid_psf', 0)) or None if row.get('land_bid_psf') else None,
        source_urls={'excel': str(row.get('source', 'Excel import'))},
        data_confidence=str(row.get('confidence', 'medium')).strip().lower(),
        data_source=str(row.get('source', '')).strip(),
        needs_review=False,
        last_scraped=datetime.utcnow(),
    )


def _update_from_excel(existing, row: Dict[str, Any]):
    """Update existing NewLaunch from Excel row."""

    existing.developer = str(row.get('developer', '')).strip() or existing.developer

    if row.get('total_units'):
        existing.total_units = int(row['total_units'])

    if row.get('psf_min'):
        existing.indicative_psf_low = float(row['psf_min'])
    if row.get('psf_max'):
        existing.indicative_psf_high = float(row['psf_max'])
    elif row.get('psf_min'):
        existing.indicative_psf_high = float(row['psf_min'])

    if row.get('tenure'):
        existing.tenure = str(row['tenure']).strip()

    if row.get('land_bid_psf'):
        existing.land_bid_psf = float(row['land_bid_psf'])

    if row.get('confidence'):
        existing.data_confidence = str(row['confidence']).strip().lower()

    if row.get('source'):
        existing.data_source = str(row['source']).strip()

    existing.source_urls = {'excel': str(row.get('source', 'Excel import'))}
    existing.needs_review = False
    existing.updated_at = datetime.utcnow()


# =============================================================================
# CLI INTEGRATION
# =============================================================================

def main():
    """CLI entry point for Excel import."""
    import argparse
    import sys
    sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

    from app import create_app
    from models.database import db

    parser = argparse.ArgumentParser(description='Import new launches from Excel')
    parser.add_argument('file', help='Path to Excel file')
    parser.add_argument('--year', type=int, default=2026, help='Launch year (default: 2026)')
    parser.add_argument('--dry-run', action='store_true', help='Validate without importing')
    parser.add_argument('--reset', action='store_true', help='Delete existing records first')

    args = parser.parse_args()

    app = create_app()
    with app.app_context():
        stats = load_new_launches_excel(
            file_path=args.file,
            db_session=db.session,
            dry_run=args.dry_run,
            reset=args.reset,
            year=args.year
        )

        if stats['errors'] or stats['validation_errors']:
            sys.exit(1)


if __name__ == '__main__':
    main()
